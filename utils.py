def format_schedule_to_markdown_table(schedule_dict, days_week):
    """
    결과를 Markdown 형식으로 출력해주는 함수
    結果をMarkdown形式で出力してくれる関数
    """
    markdown = ""

    for term, weeks in schedule_dict.items():
        markdown += f"# {term}\n\n"

        # 테이블 헤더 생성
        markdown += "| Week |"
        for day in range(1, days_week+1):
            markdown += f" Day {day} |"
        markdown += "\n|------|"
        for day in range(1, days_week+1):
            markdown += "--------|"
        markdown += "\n"

        # 각 주차별 데이터 추가
        for week, days in weeks.items():
            row = f"| {week} |"

            # 각 요일별 데이터 처리
            for day_num in range(1, days_week+1):  # day1, day2, day3
                day_key = f"day{day_num}"
                lessons = days[day_key]

                cell_content = ""
                if lessons:
                    lesson_texts = []
                    for lesson in lessons:
                        lesson_texts.append(
                            f"[{lesson['type']}: {lesson['skill']}({lesson['group_id']}{lesson['priority_order']})] {lesson['menu_name']} ({lesson['minutes']}分)"
                        )
                    cell_content = "<br>".join(lesson_texts)

                row += f" {cell_content} |"

            markdown += row + "\n"

        markdown += "\n---\n\n"

    return markdown



import pandas as pd
from typing import Dict
import openpyxl

def convert_schedule_to_excel(
    schedule_dict: Dict, 
    days_week: int, 
    output_path: str, 
    summary_df: pd.DataFrame = None,
    summary_by_term_df: pd.DataFrame = None,
    terms: list = ["前", "中", "後"],
    row_height_per_line: float = 27,  # 한 줄당 높이
    min_column_width: int = 15,       # 최소 열 너비
    max_column_width: int = 45        # 최대 열 너비
):
    """
    스케줄 딕셔너리를 엑셀 파일로 변환
    
    Args:
        schedule_dict: 스케줄 정보가 담긴 딕셔너리
        days_week: 주당 일수
        output_path: 저장할 엑셀 파일 경로
        summary_df: 전체 요약 테이블
        summary_by_term_df: term별 요약 테이블
        terms: 기간 리스트 (기본값: ["전", "중", "후"])
        row_height_per_line: 한 줄당 행 높이 (기본값: 25)
        min_column_width: 최소 열 너비 (기본값: 15)
        max_column_width: 최대 열 너비 (기본값: 40)
    """
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for term in terms:
            if term not in schedule_dict:
                continue
                
            weeks = schedule_dict[term]
            
            # DataFrame을 위한 데이터 준비
            data = []
            for week_num, days in weeks.items():
                row = {'Week': week_num}
                for day_num in range(1, days_week + 1):
                    day_key = f'day{day_num}'
                    lessons = days[day_key]
                    
                    cell_content = '\n'.join(
                        f"[{lesson['type']}: {lesson['skill']}] {lesson['menu_name']} ({lesson['minutes']}分)"
                        for lesson in lessons
                    ) if lessons else ''
                    
                    row[f'Day {day_num}'] = cell_content
                data.append(row)
            
            df = pd.DataFrame(data)
            df.to_excel(writer, sheet_name=term, index=False)
            
            worksheet = writer.sheets[term]
            
            # 열 너비 조정
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(col)
                )
                # 최소, 최대 열 너비 제한 적용
                column_width = min(max(min_column_width, max_length + 2), max_column_width)
                worksheet.column_dimensions[chr(65 + idx)].width = column_width
                
            # 줄바꿈 설정 및 셀 높이 자동 조정
            for row in worksheet.iter_rows():
                max_lines = 1
                for cell in row:
                    if cell.value:
                        cell.alignment = openpyxl.styles.Alignment(
                            wrap_text=True,      # 자동 줄바꿈
                            vertical='top',      # 상단 정렬
                            horizontal='left'    # 좌측 정렬
                        )
                        lines = str(cell.value).count('\n') + 1
                        max_lines = max(max_lines, lines)
                
                # 행 높이 설정
                worksheet.row_dimensions[cell.row].height = max_lines * row_height_per_line

            # 헤더 스타일링
            header_fill = openpyxl.styles.PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
            header_font = openpyxl.styles.Font(bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                
            # 테두리 스타일
            thin_border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
            
            # 모든 셀에 테두리 적용
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    
        # 결과 DataFrame을 마지막 시트로 추가
        if summary_df is not None:
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # 결과 시트 스타일링
            worksheet = writer.sheets['Summary']
            
            # 열 너비 조정
            for idx, col in enumerate(summary_df.columns):
                max_length = max(
                    summary_df[col].astype(str).apply(len).max(),
                    len(col)
                )
                column_width = min(max(min_column_width, max_length + 2), max_column_width)
                worksheet.column_dimensions[chr(65 + idx)].width = column_width
            
            # 헤더 스타일링
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                
            # 모든 셀에 테두리와 정렬 적용
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = openpyxl.styles.Alignment(
                        wrap_text=True,
                        vertical='center',
                        horizontal='left'
                    )
                    
        # term별 결과 DataFrame을 마지막 시트로 추가
        if summary_by_term_df is not None:
            summary_by_term_df.to_excel(writer, sheet_name="Summary_term", index=False)

            # 결과 시트 스타일링
            worksheet = writer.sheets["Summary_term"]

            # 열 너비 조정
            for idx, col in enumerate(summary_by_term_df.columns):
                max_length = max(
                    summary_by_term_df[col].astype(str).apply(len).max(), len(col)
                )
                column_width = min(
                    max(min_column_width, max_length + 2), max_column_width
                )
                worksheet.column_dimensions[chr(65 + idx)].width = column_width

            # 헤더 스타일링
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font

            # 모든 셀에 테두리와 정렬 적용
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = openpyxl.styles.Alignment(
                        wrap_text=True, vertical="center", horizontal="left"
                    )